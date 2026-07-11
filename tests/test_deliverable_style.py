"""Tests for the shared deliverable style kit — UI polish plan Phase 4.

Covers docs/UI_AND_DELIVERABLES_POLISH_PLAN.md Phase 4: one visual system
across the Excel master's four tabs and the Word reports, keyed to the app's
own brick-red light theme.

Unit-level per the plan's acceptance ("reload the workbook bytes with openpyxl
and assert freeze panes, gridlines False, header fill B3554D"; "assert docx
base font / heading colour"). Pure stdlib ``unittest``; no Streamlit run, no app
launch. Run:

    python -m unittest tests.test_deliverable_style
"""

import io
import os
import sys
import unittest
from types import SimpleNamespace

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import app  # noqa: E402  (path shim must precede the import)

# Palette from .streamlit/config.toml (light) — the single source of truth the
# whole tray is meant to share.
HEADER = "B3554D"
SUB = "DDDAD3"
LABEL = "E9E7E2"
BORDER = "C6C2B9"


def _rgb(color):
    return (color.rgb or "").upper()


class KitPaletteTests(unittest.TestCase):
    def test_kit_colours_match_theme(self):
        k = app._xl_style_kit()
        self.assertTrue(_rgb(k.header_fill.fgColor).endswith(HEADER))
        self.assertTrue(_rgb(k.sub_fill.fgColor).endswith(SUB))
        self.assertTrue(_rgb(k.label_fill.fgColor).endswith(LABEL))
        self.assertTrue(_rgb(k.border.left.color).endswith(BORDER))

    def test_kit_fonts_are_arial(self):
        k = app._xl_style_kit()
        for font in (k.header_font, k.sub_font, k.base_font, k.bold_font):
            self.assertEqual(font.name, "Arial")
        self.assertTrue(k.header_font.bold)
        self.assertTrue(_rgb(k.header_font.color).endswith("FFFFFF"))


class HelperTests(unittest.TestCase):
    """The two helpers tabs 1-3 lean on, verified through a round-trip so the
    styling is proven to survive save/reload (openpyxl bytes)."""

    def _roundtrip(self, wb):
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from openpyxl import load_workbook
        return load_workbook(buf)

    def test_header_row_paints_brick_band(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for c in range(1, 5):
            ws.cell(row=1, column=c, value=f"H{c}")
        k = app._xl_style_kit()
        app._style_header_row(ws, 1, 4, k)
        ws2 = self._roundtrip(wb).active
        for c in range(1, 5):
            cell = ws2.cell(row=1, column=c)
            self.assertTrue(_rgb(cell.fill.fgColor).endswith(HEADER))
            self.assertTrue(cell.font.bold)

    def test_finish_sheet_freeze_gridlines_widths(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        app._finish_sheet(ws, freeze="C2", widths=[12, 26, 0, 7])
        rt = self._roundtrip(wb).active
        self.assertFalse(rt.sheet_view.showGridLines)
        self.assertEqual(rt.freeze_panes, "C2")
        self.assertEqual(rt.column_dimensions["A"].width, 12)
        self.assertEqual(rt.column_dimensions["B"].width, 26)
        # A 0/None width entry is skipped; a later real width still lands.
        self.assertEqual(rt.column_dimensions["D"].width, 7)


class ClassroomEntryStyleTests(unittest.TestCase):
    """Tab 4 restyled from navy to brick — same structure, new colours — and
    the paste-back data (names, ids, order) unchanged."""

    def setUp(self):
        from openpyxl import Workbook
        self.Workbook = Workbook
        self._orig_st = app.st
        self._orig_names = app._classroom_folder_assignment_names
        self._orig_scores = app.all_scores
        roster = [
            {"key": "1", "name": "Aoki Daichi", "first": "Daichi"},
            {"key": "2", "name": "Baba Chiaki", "first": "Chiaki"},
        ]
        self.ss = {"roster": roster}
        app.st = SimpleNamespace(session_state=self.ss)
        self.students = [SimpleNamespace(student_id=e["key"], name=e["name"])
                         for e in roster]
        app._classroom_folder_assignment_names = lambda: []
        app.all_scores = lambda: []

    def tearDown(self):
        app.st = self._orig_st
        app._classroom_folder_assignment_names = self._orig_names
        app.all_scores = self._orig_scores

    def _build(self):
        wb = self.Workbook()
        app._append_classroom_entry_sheet(wb, list(self.students))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from openpyxl import load_workbook
        return load_workbook(buf)["Classroom Entry"]

    def test_header_is_brick_not_navy(self):
        ws = self._build()
        fill = _rgb(ws.cell(row=1, column=1).fill.fgColor)
        self.assertTrue(fill.endswith(HEADER))
        self.assertFalse(fill.endswith("1F4E78"))   # the old navy is gone

    def test_freeze_and_gridlines_intact(self):
        ws = self._build()
        self.assertEqual(ws.freeze_panes, "C3")
        self.assertFalse(ws.sheet_view.showGridLines)

    def test_paste_back_data_unchanged(self):
        ws = self._build()
        # Fixed headers and per-student rows keep their exact values/positions.
        self.assertEqual(ws.cell(row=1, column=1).value, "Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Student ID")
        # Rows keep Latin first-name order (Chiaki < Daichi), ids intact.
        self.assertEqual(ws.cell(row=3, column=2).value, "2")   # Baba Chiaki
        self.assertEqual(ws.cell(row=4, column=2).value, "1")   # Aoki Daichi


class ReportDocStyleTests(unittest.TestCase):
    def test_apply_report_styles(self):
        from docx import Document
        from docx.shared import RGBColor
        doc = Document()
        app._apply_report_styles(doc)
        normal = doc.styles["Normal"]
        self.assertEqual(normal.font.name, "Arial")
        self.assertEqual(normal.font.color.rgb, RGBColor(0x38, 0x35, 0x2F))
        h1 = doc.styles["Heading 1"]
        self.assertEqual(h1.font.color.rgb, RGBColor(0xB3, 0x55, 0x4D))
        self.assertTrue(h1.font.bold)
        h2 = doc.styles["Heading 2"]
        self.assertEqual(h2.font.color.rgb, RGBColor(0x9C, 0x4A, 0x43))

    def test_h1_has_bottom_border_applied_once(self):
        from docx import Document
        from docx.oxml.ns import qn
        doc = Document()
        app._apply_report_styles(doc)
        app._apply_report_styles(doc)   # idempotent — no duplicate rule
        pPr = doc.styles["Heading 1"].element.find(qn("w:pPr"))
        self.assertIsNotNone(pPr)
        self.assertEqual(len(pPr.findall(qn("w:pBdr"))), 1)

    def test_new_report_document_is_prestyled(self):
        # _new_report_document must route through _apply_report_styles so every
        # export (pack, single, mail-merge, class comments) inherits the look.
        doc = app._new_report_document()
        self.assertEqual(doc.styles["Normal"].font.name, "Arial")


if __name__ == "__main__":
    unittest.main()
